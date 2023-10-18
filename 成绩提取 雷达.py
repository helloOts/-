import tkinter as tk
from tkinter import filedialog
import os
import matplotlib.pyplot as plt
import pandas as pd
from math import pi
import openpyxl
import numpy as np

plt.rcParams['font.sans-serif'] = ['Arial Unicode MS']  # 使用SimHei字体
plt.rcParams['axes.unicode_minus'] = False  # 解决保存图像时'-'显示为方块的问题

# 使用tkinter的filedialog让用户选择一个文件夹
root = tk.Tk()
root.withdraw()  # 隐藏主窗口
folder_path = filedialog.askdirectory(title="请选择文件夹")

for file in os.listdir(folder_path):
    if file.endswith(".xlsx"):
        file_path = os.path.join(folder_path, file)
        
        # 创建一个与Excel文件名相同的子文件夹来保存图片
        output_folder = os.path.join(folder_path, os.path.splitext(file)[0])
        if not os.path.exists(output_folder):
            os.makedirs(output_folder)
        
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        sheet_names = workbook.sheetnames
        for sheet_name in sheet_names:
            if sheet_name=="雷达图":
                sheet = workbook[sheet_name]
                for row in sheet.iter_rows(min_row=3, values_only=True): 
                    name = row[0]
                    tiaosheng= row[1]
                    tiaosheng1= row[7]

                    yangwoqizuo = row[2]
                    yangwoqizuo1 = row[8]
                    zwtqq = row[3]
                    zwtqq1 = row[9]
                    pao = row[4]
                    pao1 = row[10]
                    lidingtiaoyuan = row[5]
                    lidingtiaoyuan1 = row[11]


                    # 在雷达图绘制部分之前，创建一个新的列表来存储具体的成绩数据
                    actual_scores = [tiaosheng1, yangwoqizuo1, zwtqq1, pao1, lidingtiaoyuan1]
                    actual_scores += actual_scores[:1]  # 将第一个值放到最后，以封闭图形



                    # 创建数据框
                    df = pd.DataFrame({
                        'name': [name],
                        '跳绳': [tiaosheng],
                        '仰卧起坐': [yangwoqizuo],
                        '坐位体前屈': [zwtqq],
                        '50M': [pao],
                        '立定跳远': [lidingtiaoyuan],
                    })
                    #print(name,tiaosheng,yangwoqizuo,zwtqq,pao,lidingtiaoyuan)
                    # 变量类别
                    categories = list(df)[1:]
                    # 变量类别个数
                    N = len(categories)

                    # 绘制数据的第一行
                    values = df.loc[0].drop('name').values.flatten().tolist()
                    # 将第一个值放到最后，以封闭图形
                    values += values[:1]

                    # 设置每个点的角度值
                    angles = [n / float(N) * 2 * pi for n in range(N)]
                    angles += angles[:1]

                    # 定义五种鲜艳的颜色，用于填充雷达图的部分
                    colors = ['#FF9999', '#66B2FF', '#99FF99', '#FFCC99', '#FFD700']

                    # 定义五种鲜艳的颜色，用于标签
                    label_colors = ['#FF9999', '#66B2FF', '#99FF99', '#FFCC99', '#FFD700']

                    # 初始化一个极坐标网格
                    ax = plt.subplot(111, polar=True, facecolor='#E6F2F8')

                    # 设置x轴的标签，并将标签旋转50度以避免重叠
                    xticks = plt.xticks(angles[:-1], categories, size=10, fontweight='bold', rotation=50)

                    # 遍历每个标签并设置对应的颜色
                    for i, label in enumerate(xticks[1]):
                        label.set(color=label_colors[i])

                    # 设置y轴的标签位置和标签值
                    plt.yticks([100], [""], color="purple", size=10)
                    plt.ylim(0, 110)

                    # 在极坐标上绘制数据
                    ax.plot(angles, values, linewidth=2, linestyle='dashed', color='orange')

                    # 为雷达图的每个部分填充颜色
                    for i in range(N):
                        ax.fill(angles[i:i+2] + [0], values[i:i+2] + [0], color=colors[i], alpha=0.7)

                    #为每个数据点添加注释，显示具体的值，并为注释添加背景颜色
                    # ... [其他代码保持不变]

#为每个数据点添加注释，显示具体的值，并为注释添加背景颜色
                    # ... [其他代码保持不变]

#为每个数据点添加注释，显示具体的值，并为注释添加背景颜色
                   # ... [其他代码保持不变]

#为每个数据点添加注释，显示具体的值，并为注释添加背景颜色
                    # ... [其他代码保持不变]

#为每个数据点添加注释，显示具体的值，并为注释添加背景颜色
                    for i, (value, actual_score) in enumerate(zip(values[:-1], actual_scores[:-1])):
                        category = categories[i]  # 获取当前数据点的类别
                        if value == "#N/A":
                            value_num = 0  # 或者您可以选择其他默认值
                        else:
                            value_num = float(value)
                        
                        # 如果是50米成绩，保留小数点
                        if category == "50M":
                            fraction_score = f"{actual_score:.2f}/{value_num:.2f}"  # 保留两位小数
                        else:
                            fraction_score = f"{int(actual_score)}/{int(value_num)}"
                        
                        # 在雷达图上添加注释
                        # 左边的成绩
                        annotation_left = ax.annotate(f"{actual_score}", 
                                                    xy=(angles[i], value_num), 
                                                    xytext=(0, 0),  # 调整注释位置
                                                    textcoords='offset points',
                                                    ha="center", va="center", 
                                                    color="white", size=9, fontweight='bold')
                        annotation_left.set_bbox(dict(facecolor=colors[i], alpha=0.9, edgecolor='none'))
                        
                        # 右边的分数
                        annotation_right = ax.annotate(f"{value_num}", 
                                                    xy=(angles[i], value_num), 
                                                    xytext=(0, -20),  # 调整注释位置
                                                    textcoords='offset points',
                                                    ha="center", va="center", 
                                                    color="white", size=11, fontweight='bold', backgroundcolor='red')  # 使用醒目的黄色背景
                        annotation_right.set_bbox(dict(facecolor='#ed9408', alpha=0.9, edgecolor='none'))

# ... [其他代码保持不变]



# ... [其他代码保持不变]

                       


                    # 设置图表标题
                    plt.title(name, size=16, color="white", y=1.07, fontweight='bold', bbox=dict(facecolor='red', alpha=0.8, edgecolor='#FF4500', boxstyle='round,pad=0.3'))


                    # 在雷达图的中心添加一个注释，显示总分，并为注释添加背景颜色
                    numeric_values = [float(v) if v != "#N/A" else 0 for v in values[:-1]]
                    score_annotation = ax.annotate(f"{sum(numeric_values)}", 

                                                xy=(0, 0),
                                                ha="center", va="center", 
                                                color="black", size=15, fontweight='bold', style='italic')
                    score_annotation.set_bbox(dict(facecolor='#FFD700', alpha=0.8, edgecolor='#FF4500', boxstyle='round,pad=0.3'))
                    
                    # Save the plot
                    plt.savefig(os.path.join(output_folder, f'{name}.png'), dpi=300, bbox_inches='tight')
                    plt.close()




